from fistf_ko_engine import build_knockout_sheet

# -*- coding: utf-8 -*-
"""
Torneo_FISTF_v4d_FINAL_FIXED.py
"""
import os
import time
import traceback
import sys
from datetime import datetime
from typing import List, Dict, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.worksheet import Worksheet

# ------------------------------
# Configurazione da variabili d'ambiente
# ------------------------------
def get_env_bool(name, default=False):
    val = os.environ.get(name, str(default)).lower()
    return val in ('true', '1', 'yes', 'y')

DEBUG_LOG = get_env_bool('FISTF_DEBUG', True)
PLACEHOLDER_MODE = get_env_bool('FISTF_PLACEHOLDER_MODE', True)

DEFAULT_CAMPI = 2
RANDOM_SEED = 42
CLASSIFICA_SIDE_START_COL = 12
TIEBREAK_ORDER_NOTE = "P → HTH_P → HTH_DG → HTH_GF → DG → GF → (Shoot-out operativo)"
MAX_GROUPS_PER_CATEGORY = 16

def log_debug(msg):
    """Funzione di logging sicura che non dà errori se stdout è None"""
    if DEBUG_LOG:
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        try:
            # Prova a stampare su stdout
            if sys.stdout is not None:
                print(f"[ENGINE {timestamp}] {msg}")
                sys.stdout.flush()
            else:
                # Se stdout è None, prova a scrivere su stderr
                if sys.stderr is not None:
                    print(f"[ENGINE {timestamp}] {msg}", file=sys.stderr)
                    sys.stderr.flush()
        except (AttributeError, IOError):
            # Se anche quello fallisce, ignora silenziosamente
            pass

# ------------------------------
# Utilities
# ------------------------------
def _log(m: str):
    """Log semplice senza timestamp"""
    try:
        if sys.stdout is not None:
            print(f"[FISTF v4d] {m}")
    except (AttributeError, IOError):
        pass

def clean_cell(val):
    if val is None:
        return None
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return None
    if hasattr(val, 'item'):
        try:
            val = val.item()
        except Exception:
            pass
    if isinstance(val, str):
        val = ILLEGAL_CHARACTERS_RE.sub('', val)
        val = val.replace('\\', '')
    return val

def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in out.columns:
        out[c] = out[c].map(clean_cell)
    return out

def safe_save_workbook(wb: Workbook, final_path: str, tries: int = 3, sleep_sec: float = 0.8) -> str:
    """Salva workbook con gestione file bloccati"""
    folder = os.path.dirname(final_path) or os.getcwd()
    base = os.path.splitext(os.path.basename(final_path))[0]
    ext = os.path.splitext(final_path)[1] or ".xlsx"
    os.makedirs(folder, exist_ok=True)
    
    tmp = os.path.join(folder, f"~${base}_{int(time.time()*1000)}.tmp")
    
    try:
        wb.save(tmp)
        log_debug(f"   💾 Temporaneo: {tmp}")
        
        for attempt in range(tries):
            try:
                if os.path.exists(final_path):
                    os.remove(final_path)
                os.rename(tmp, final_path)
                log_debug(f"   ✅ Salvato: {final_path}")
                return final_path
            except PermissionError:
                log_debug(f"   ⚠️ Tentativo {attempt + 1}/{tries}: file bloccato, aspetto...")
                time.sleep(sleep_sec)
        
        alt_path = os.path.join(folder, f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}")
        os.rename(tmp, alt_path)
        log_debug(f"   ✅ Salvato come: {alt_path}")
        return alt_path
        
    except Exception as e:
        log_debug(f"   ❌ Errore salvataggio: {e}")
        fallback = os.path.join(folder, f"{base}_FALLBACK.xlsx")
        wb2 = Workbook()
        ws = wb2.active
        ws.title = 'ERRORE'
        ws.append(["Errore durante il salvataggio", str(e)])
        for line in traceback.format_exc().splitlines():
            ws.append([line])
        wb2.save(fallback)
        return fallback
        
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except:
            pass

# ------------------------------
# Input lettura
# ------------------------------
def read_campi_mapping_if_any(xls: pd.ExcelFile, preset: Dict[str, int]) -> Dict[str, int]:
    m = dict(preset)
    try:
        if 'Campi' in xls.sheet_names:
            df = pd.read_excel(xls, 'Campi', engine='openpyxl')
            cols = {str(c).strip().upper(): c for c in df.columns}
            cat = cols.get('CATEGORIA') or cols.get('CATEGORY')
            n = cols.get('CAMPI') or cols.get('FIELDS') or cols.get('COURTS')
            if cat and n:
                for _, r in df.iterrows():
                    k = str(r.get(cat, '')).strip()
                    try:
                        v = int(r.get(n, DEFAULT_CAMPI))
                        if k:
                            m[k] = max(1, v)
                    except Exception:
                        pass
            _log(f"Mapping campi: {m}")
        else:
            _log("Foglio 'Campi' non presente: uso default per tutte.")
    except Exception as e:
        _log(f"Errore lettura 'Campi': {e}. Uso default/preset.")
    return m

def load_players_sheet(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(xls, sheet_name=sheet_name, engine='openpyxl')
    df.columns = [str(c).strip().upper() for c in df.columns]
    
    if 'U' in df.columns and not any(c in df.columns for c in ('GIOCATORE', 'NOME', 'PLAYER')):
        df = df.rename(columns={'U': 'GIOCATORE'})
    
    col_map = {}
    mapping = {
        'NOME': ('GIOCATORE', 'NOME', 'PLAYER'),
        'CLUB': ('CLUB',),
        'NAZIONE': ('NAZIONE', 'NATION', 'COUNTRY'),
        'POS': ('POS', 'RANK', 'WR_POS', 'WR')
    }
    
    for key, cands in mapping.items():
        found = None
        for c in df.columns:
            for cand in cands:
                if cand == c or cand in c:
                    found = c
                    break
            if found:
                break
        if found:
            col_map[key] = found
    
    req = {'NOME', 'CLUB', 'POS'}
    if not req.issubset(set(col_map.keys())):
        raise ValueError(f"[{sheet_name}] Colonne mancanti: {req - set(col_map.keys())}")
    
    out = pd.DataFrame({
        'NOME': df[col_map['NOME']].astype(str).str.strip(),
        'CLUB': df[col_map['CLUB']].astype(str).str.strip(),
        'POS': pd.to_numeric(df[col_map['POS']], errors='coerce'),
    })
    out['NAZIONE'] = df[col_map['NAZIONE']].astype(str).str.strip() if 'NAZIONE' in col_map else ''
    out = out[out['NOME'].str.len() > 0]
    out = out[out['POS'].notna()].copy()
    out['POS'] = out['POS'].astype(int)
    return out.sort_values('POS').reset_index(drop=True)

# ------------------------------
# Group size logic
# ------------------------------
def compute_group_sizes(N: int) -> List[int]:
    if N <= 0:
        return []
    q, r = divmod(N, 4)
    if r == 0:
        return [4] * q
    if r == 1:
        if q >= 1:
            return [5] + [4] * (q - 1)
        else:
            return [5]
    if r == 2:
        if q >= 2:
            return [5, 5] + [4] * (q - 2)
        else:
            return [3, 3]
    return [3] + [4] * q

def make_group_labels(G: int) -> List[str]:
    import string
    base = list(string.ascii_uppercase)
    if G <= len(base):
        return base[:G]
    labels = base[:]
    k = 1
    while len(labels) < G:
        for b in base:
            labels.append(f"{b}{k}")
            if len(labels) == G:
                break
        k += 1
    return labels

# ------------------------------
# Draw con vincoli
# ------------------------------
def build_pots(players: pd.DataFrame, G: int) -> List[pd.DataFrame]:
    return [players.iloc[i:i+G].reset_index(drop=True) for i in range(0, len(players), G)]

def draw_groups_fistf(players: pd.DataFrame, sizes: List[int], random_seed=RANDOM_SEED,
                       avoid_fields=("CLUB", "NAZIONE")) -> Dict[str, List[dict]]:
    rng = np.random.default_rng(random_seed)
    labels = make_group_labels(len(sizes))
    groups = {L: [] for L in labels}
    players_sorted = players.sort_values('POS').reset_index(drop=True)
    G = len(sizes)

    seeds = players_sorted.iloc[:G]
    for L, (_, row) in zip(labels, seeds.iterrows()):
        groups[L].append(row.to_dict())

    def conflict_score(L: str, prow: dict) -> int:
        s = 0
        for f in avoid_fields:
            val = str(prow.get(f, '')).upper()
            if not val:
                continue
            s += sum(1 for x in groups[L] if str(x.get(f, '')).upper() == val)
        return s

    rest = players_sorted.iloc[G:]
    pot2 = rest.iloc[:G]
    rest2 = rest.iloc[G:]

    gi = list(range(G))
    rng.shuffle(gi)
    for _, prow in pot2.iterrows():
        eligible = [labels[g] for g in gi if len(groups[labels[g]]) < sizes[g] and len(groups[labels[g]]) == 1]
        if not eligible:
            eligible = [L for i, L in enumerate(labels) if len(groups[L]) < sizes[i]]
        chosen = sorted([(conflict_score(L, prow.to_dict()), rng.random(), L) for L in eligible])[0][2]
        groups[chosen].append(prow.to_dict())

    for pot in build_pots(rest2, G):
        for _, prow in pot.iterrows():
            cands = [L for i, L in enumerate(labels) if len(groups[L]) < sizes[i]]
            chosen = sorted([(conflict_score(L, prow.to_dict()), rng.random(), L) for L in cands])[0][2]
            groups[chosen].append(prow.to_dict())

    return groups

# ------------------------------
# Fixtures
# ------------------------------
def round_robin(names: List[str]) -> List[List[Tuple[str, str]]]:
    n = len(names)
    arr = names[:]
    if n % 2 == 1:
        arr.append(None)
        n += 1
    half = n // 2
    out = []
    for _ in range(n - 1):
        pairs = []
        for i in range(half):
            a, b = arr[i], arr[n - 1 - i]
            if a and b:
                pairs.append((a, b))
        arr = [arr[0]] + [arr[-1]] + arr[1:-1]
        out.append(pairs)
    return out

def reorder_first_match_if_conflict(rr: List[List[Tuple[str, str]]], players_meta: pd.DataFrame,
                                    avoid_fields=("CLUB", "NAZIONE")) -> List[List[Tuple[str, str]]]:
    if not rr:
        return rr
    
    club_of = dict(zip(players_meta['Giocatore'], players_meta['Club']))
    naz_of = dict(zip(players_meta['Giocatore'], players_meta['Nazione']))
    players = players_meta['Giocatore'].tolist()

    def is_conflict(a, b):
        return (club_of.get(a, '') and club_of.get(a, '') == club_of.get(b, '')) or \
               (naz_of.get(a, '') and naz_of.get(a, '') == naz_of.get(b, ''))

    conflicts = []
    for i in range(len(players)):
        for j in range(i+1, len(players)):
            a, b = players[i], players[j]
            if is_conflict(a, b):
                conflicts.append((a, b))

    if not conflicts:
        return rr

    first_round = rr[0][:]
    present = any({tuple(sorted(m)) for m in first_round} & {tuple(sorted(c)) for c in conflicts})
    if present:
        return rr

    for r_ix in range(1, len(rr)):
        for m_ix, (a, b) in enumerate(rr[r_ix]):
            if tuple(sorted((a, b))) in {tuple(sorted(c)) for c in conflicts}:
                if rr[0]:
                    rr[r_ix][m_ix], rr[0][0] = rr[0][0], rr[r_ix][m_ix]
                    return rr
    return rr

def build_fixtures_with_fields_and_slots(groups: dict, num_campi: int, df_gironi: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for L, plist in groups.items():
        names = [p['NOME'] for p in plist]
        if len(names) < 2:
            continue
        rr = round_robin(names)
        meta = df_gironi[df_gironi['Girone'] == L][['Giocatore', 'Club', 'Nazione']].drop_duplicates()
        rr = reorder_first_match_if_conflict(rr, meta)
        for r_ix, matches in enumerate(rr, start=1):
            for m_ix, (a, b) in enumerate(matches, start=1):
                rows.append({'Girone': L, 'Giornata': r_ix, 'Match #': m_ix, 'Casa': a, 'Trasferta': b})

    if not rows:
        return pd.DataFrame(columns=['Girone', 'Giornata', 'Match #', 'Casa', 'Trasferta', 'Campo', 'Slot',
                                     'Gol Casa', 'Gol Trasferta', 'Arbitro'])

    df = pd.DataFrame(rows).sort_values(['Giornata', 'Girone', 'Match #']).reset_index(drop=True)
    df['Campo'] = None
    df['Slot'] = None

    slot_map = []
    for giornata, sub in df.groupby('Giornata', sort=True):
        idxs = sub.index.tolist()
        for i, ridx in enumerate(idxs):
            slot = (i // num_campi) + 1
            campo = (i % num_campi) + 1
            slot_map.append((ridx, slot, campo))
    
    for ridx, slot, campo in slot_map:
        df.at[ridx, 'Slot'] = slot
        df.at[ridx, 'Campo'] = campo

    def _si(s):
        return pd.to_numeric(s, errors='coerce').fillna(10**9).astype(int)

    df['_Giornata_'] = _si(df['Giornata'])
    df['_Slot_'] = _si(df['Slot'])
    df['_Campo_'] = _si(df['Campo'])
    df = df.sort_values(['_Giornata_', '_Slot_', '_Campo_', 'Girone', 'Match #']) \
           .drop(columns=['_Giornata_', '_Slot_', '_Campo_']) \
           .reset_index(drop=True)
    df['Gol Casa'] = ''
    df['Gol Trasferta'] = ''
    df['Arbitro'] = None
    return df

# ------------------------------
# Referees
# ------------------------------
def assign_referees_by_slot(df_cal: pd.DataFrame, df_gironi: pd.DataFrame, category_name: str = None,
                            excluded_categories=("U12", "Beginners")) -> pd.DataFrame:
    df = df_cal.copy()
    if df.empty:
        df['Arbitro'] = "-"
        return df

    if category_name and category_name.strip().lower() in tuple(x.lower() for x in excluded_categories):
        df['Arbitro'] = "-"
        return df

    df['Arbitro'] = None
    meta = df_gironi[["Giocatore", "Club", "Nazione"]].drop_duplicates()
    club_of = dict(zip(meta['Giocatore'], meta['Club']))
    naz_of = dict(zip(meta['Giocatore'], meta['Nazione']))
    players = meta['Giocatore'].unique().tolist()
    ref_count = {p: 0 for p in players}

    playing = {(g, s): set(sub['Casa']).union(set(sub['Trasferta'])) 
               for (g, s), sub in df.groupby(['Giornata', 'Slot'])}
    used = {(g, s): set() for g, s in df[['Giornata', 'Slot']].drop_duplicates().itertuples(index=False, name=None)}

    for (giorn, slot), sub in df.groupby(['Giornata', 'Slot'], sort=True):
        p_play = playing[(giorn, slot)]
        for ridx, row in sub.sort_values('Campo').iterrows():
            casa, trasf = row['Casa'], row['Trasferta']
            block_club = {club_of.get(casa, ''), club_of.get(trasf, '')}
            block_naz = {naz_of.get(casa, ''), naz_of.get(trasf, '')}

            cand = [p for p in players if p not in p_play and p not in used[(giorn, slot)] 
                    and p != casa and p != trasf and
                    club_of.get(p, '') not in block_club and naz_of.get(p, '') not in block_naz]
            
            if not cand:
                cand = [p for p in players if p not in p_play and p not in used[(giorn, slot)] 
                        and p != casa and p != trasf and club_of.get(p, '') not in block_club]
            
            if not cand:
                cand = [p for p in players if p not in p_play and p not in used[(giorn, slot)] 
                        and p != casa and p != trasf]
            
            if not cand:
                cand = ["-"]

            chosen = sorted(cand, key=lambda x: ref_count.get(x, 10**9) if x != "-" else 10**12)[0]
            df.at[ridx, 'Arbitro'] = chosen
            if chosen != "-":
                ref_count[chosen] += 1
                used[(giorn, slot)].add(chosen)
    return df

# ------------------------------
# Conformity report
# ------------------------------
def compute_draw_conformity(df_gironi: pd.DataFrame) -> pd.DataFrame:
    if df_gironi is None or df_gironi.empty:
        return pd.DataFrame([{
            'Girone': '-', 'Conflitti CLUB': 0, 'Min CLUB (cat)': 0,
            'Conflitti NAZIONE': 0, 'Min NAZIONE (cat)': 0,
            'Raggiunti min teorici?': 'Sì'
        }])

    gruppi = df_gironi['Girone'].dropna().unique().tolist()
    G = len(gruppi)
    club_counts = df_gironi['Club'].astype(str).str.upper().value_counts()
    naz_counts = df_gironi['Nazione'].astype(str).str.upper().value_counts()
    min_club = int(sum(max(0, int(c) - G) for c in club_counts))
    min_naz = int(sum(max(0, int(c) - G) for c in naz_counts))

    out = []
    for L in gruppi:
        sub = df_gironi[df_gironi['Girone'] == L]
        cc = sub['Club'].astype(str).str.upper().value_counts()
        nc = sub['Nazione'].astype(str).str.upper().value_counts()
        confl_club = int(sum(v - 1 for v in cc if v > 1))
        confl_naz = int(sum(v - 1 for v in nc if v > 1))
        out.append({'Girone': L, 'Conflitti CLUB': confl_club, 'Min CLUB (cat)': min_club,
                    'Conflitti NAZIONE': confl_naz, 'Min NAZIONE (cat)': min_naz})
    
    df = pd.DataFrame(out)
    ok_club = int(df['Conflitti CLUB'].sum()) == min_club
    ok_naz = int(df['Conflitti NAZIONE'].sum()) == min_naz
    df['Raggiunti min teorici?'] = 'Sì' if (ok_club and ok_naz) else 'No'
    return df

# ------------------------------
# Excel writing
# ------------------------------
def write_df_at(ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    df = clean_df(df)
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    r = start_row + 1
    for row in df.itertuples(index=False):
        for j, v in enumerate(row, start=start_col):
            ws.cell(row=r, column=j, value=clean_cell(v))
        r += 1
    for j, col in enumerate(df.columns, start=start_col):
        try:
            mx = max(len(str(col)), *(len(str(x)) for x in df[col].astype(str)))
        except Exception:
            mx = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(mx + 2, 60)
    return r

def write_calendar_and_classifica_sheet(wb: Workbook, cat: str, calendario: pd.DataFrame, classifica_seed: pd.DataFrame):
    sh = cat[:31]
    if sh in wb.sheetnames:
        del wb[sh]
    ws = wb.create_sheet(sh)

    ws.cell(row=1, column=1, value=f"Calendario {cat}")
    ws.cell(row=1, column=CLASSIFICA_SIDE_START_COL, value=f"Classifica {cat}")
    ws.cell(row=2, column=11, value="KO")

    cal_cols = ['Girone', 'Giornata', 'Match #', 'Casa', 'Trasferta', 'Campo', 'Slot', 'Gol Casa', 'Gol Trasferta', 'Arbitro']
    r_after = write_df_at(ws, calendario[cal_cols], start_row=2, start_col=1)
    cal_first = 3
    cal_last = r_after - 1

    cls = CLASSIFICA_SIDE_START_COL
    headers = ['Girone', 'Pos', 'Giocatore', 'Punti', 'GF', 'GS', 'DG', 'HTH_P', 'HTH_GF', 'HTH_GS', 'HTH_DG']
    for j, h in enumerate(headers, start=cls):
        ws.cell(row=2, column=j, value=h)

    start = 3
    n_rows = len(classifica_seed) if classifica_seed is not None else 0
    last = start + n_rows - 1
    if n_rows <= 0:
        for j, h in enumerate(headers, start=cls):
            ws.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 12)
        ws.column_dimensions['K'].width = 10
        return

    A = f"$A${cal_first}:$A${cal_last}"
    D = f"$D${cal_first}:$D${cal_last}"
    E = f"$E${cal_first}:$E${cal_last}"
    H = f"$H${cal_first}:$H${cal_last}"
    I = f"$I${cal_first}:$I${cal_last}"

    colL = get_column_letter(cls)
    colM = get_column_letter(cls + 1)
    colN = get_column_letter(cls + 2)
    colO = get_column_letter(cls + 3)

    Lrng = f"${colL}$3:${colL}${last}"
    Nrng = f"${colN}$3:${colN}${last}"
    Orng = f"${colO}$3:${colO}${last}"

    for i, (girone, player) in enumerate(classifica_seed[['Girone', 'Giocatore']].itertuples(index=False), start=start):
        ws.cell(row=i, column=cls + 0, value=str(girone))
        ws.cell(row=i, column=cls + 2, value=str(player))

        cg   = f"$" + get_column_letter(cls)     + f"${i}"
        cp   = f"$" + get_column_letter(cls + 2) + f"${i}"
        cpts = f"$" + get_column_letter(cls + 3) + f"${i}"

        f_P = ("=IFERROR("
               f"3*SUMPRODUCT(({A}={cg})*({D}={cp})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({H}>{I}))"
               f"+1*SUMPRODUCT(({A}={cg})*({D}={cp})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({H}={I}))"
               f"+3*SUMPRODUCT(({A}={cg})*({E}={cp})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({I}>{H}))"
               f"+1*SUMPRODUCT(({A}={cg})*({E}={cp})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({I}={H})),0)")
        ws.cell(row=i, column=cls + 3, value=f_P)

        f_GF = f"=IFERROR(SUMPRODUCT(({A}={cg})*({D}={cp})*(--ISNUMBER({H}))*{H})+SUMPRODUCT(({A}={cg})*({E}={cp})*(--ISNUMBER({I}))*{I}),0)"
        f_GS = f"=IFERROR(SUMPRODUCT(({A}={cg})*({D}={cp})*(--ISNUMBER({I}))*{I})+SUMPRODUCT(({A}={cg})*({E}={cp})*(--ISNUMBER({H}))*{H}),0)"
        ws.cell(row=i, column=cls + 4, value=f_GF)
        ws.cell(row=i, column=cls + 5, value=f_GS)
        ws.cell(row=i, column=cls + 6, value=f"={get_column_letter(cls+4)}{i}-{get_column_letter(cls+5)}{i}")

        f_HTH_GF = ("=IFERROR("
                    f"SUMPRODUCT(({A}={cg})*({D}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{E})={cpts})*(--ISNUMBER({H}))*{H})+"
                    f"SUMPRODUCT(({A}={cg})*({E}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{D})={cpts})*(--ISNUMBER({I}))*{I}),0)")
        f_HTH_GS = ("=IFERROR("
                    f"SUMPRODUCT(({A}={cg})*({D}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{E})={cpts})*(--ISNUMBER({I}))*{I})+"
                    f"SUMPRODUCT(({A}={cg})*({E}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{D})={cpts})*(--ISNUMBER({H}))*{H}),0)")
        f_HTH_P = ("=IFERROR("
                    f"3*SUMPRODUCT(({A}={cg})*({D}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{E})={cpts})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({H}>{I}))"
                    f"+1*SUMPRODUCT(({A}={cg})*({D}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{E})={cpts})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({H}={I}))"
                    f"+3*SUMPRODUCT(({A}={cg})*({E}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{D})={cpts})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({I}>{H}))"
                    f"+1*SUMPRODUCT(({A}={cg})*({E}={cp})*(SUMIFS({Orng},{Lrng},{cg},{Nrng},{D})={cpts})*(--ISNUMBER({H}))*(--ISNUMBER({I}))*({I}={H})),0)")
        ws.cell(row=i, column=cls + 7, value=f_HTH_P)
        ws.cell(row=i, column=cls + 8, value=f_HTH_GF)
        ws.cell(row=i, column=cls + 9, value=f_HTH_GS)
        ws.cell(row=i, column=cls + 10, value=f"={get_column_letter(cls+8)}{i}-{get_column_letter(cls+9)}{i}")

        rng_g     = f"$" + get_column_letter(cls)     + f"$3:$" + get_column_letter(cls)     + f"${last}"
        rng_p     = f"$" + get_column_letter(cls + 3) + f"$3:$" + get_column_letter(cls + 3) + f"${last}"
        rng_hthp  = f"$" + get_column_letter(cls + 7) + f"$3:$" + get_column_letter(cls + 7) + f"${last}"
        rng_hthdg = f"$" + get_column_letter(cls + 10)+ f"$3:$" + get_column_letter(cls + 10)+ f"${last}"
        rng_hthgf = f"$" + get_column_letter(cls + 8) + f"$3:$" + get_column_letter(cls + 8) + f"${last}"
        rng_dg    = f"$" + get_column_letter(cls + 6) + f"$3:$" + get_column_letter(cls + 6) + f"${last}"
        rng_gf    = f"$" + get_column_letter(cls + 4) + f"$3:$" + get_column_letter(cls + 4) + f"${last}"

        cp_i   = f"$" + get_column_letter(cls + 3) + f"${i}"
        chp_i  = f"$" + get_column_letter(cls + 7) + f"${i}"
        chdg_i = f"$" + get_column_letter(cls + 10)+ f"${i}"
        chgf_i = f"$" + get_column_letter(cls + 8) + f"${i}"
        cdg_i  = f"$" + get_column_letter(cls + 6) + f"${i}"
        cgf_i  = f"$" + get_column_letter(cls + 4) + f"${i}"

        fPOS  = '=1'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},">"&{cp_i})'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},"="&{cp_i},{rng_hthp},">"&{chp_i})'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},"="&{cp_i},{rng_hthp},"="&{chp_i},{rng_hthdg},">"&{chdg_i})'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},"="&{cp_i},{rng_hthp},"="&{chp_i},{rng_hthdg},"="&{chdg_i},{rng_hthgf},">"&{chgf_i})'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},"="&{cp_i},{rng_hthp},"="&{chp_i},{rng_hthdg},"="&{chdg_i},{rng_hthgf},"="&{chgf_i},{rng_dg},">"&{cdg_i})'
        fPOS += f'+COUNTIFS({rng_g},{cg},{rng_p},"="&{cp_i},{rng_hthp},"="&{chp_i},{rng_hthdg},"="&{chdg_i},{rng_hthgf},"="&{chgf_i},{rng_dg},"="&{cdg_i},{rng_gf},">"&{cgf_i})'
        ws.cell(row=i, column=cls + 1, value=fPOS)

        formula_KO = f"=(CODE(UPPER(${colL}{i}))-64) & CHAR(${colM}{i}+64)"
        ws.cell(row=i, column=11, value=formula_KO)

    for j, h in enumerate(headers, start=cls):
        ws.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 12)
    ws.column_dimensions['K'].width = 10

    # Crea foglio finale con i KO
    dest_name = f"Finale_{sh}"[:31]
    try:
        build_knockout_sheet(
            wb=wb,
            source_ws_name=sh,
            dest_ws_name=dest_name,
            cls_col=CLASSIFICA_SIDE_START_COL,
            start_row=1
        )
    except Exception as e:
        log_debug(f"   ⚠️ Errore creazione foglio finale: {e}")

def export_multicat_excel(path: str, per_cat_output: Dict[str, Dict[str, pd.DataFrame]], campi_map: Dict[str, int]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Istruzioni_Globali'

    note = pd.DataFrame({
        'Parametro': [
            'Input', 'Struttura', 'Campi per categoria', 'Slot', 'Arbitri per categoria',
            'Tie-breakers', 'Cap 5-0', 'Forfeit post-2025', 'Note U12/Beginner'
        ],
        'Valore': [
            'Categorie.xlsx (un foglio per categoria)',
            'Ogni foglio: GIOCATORE/NOME, CLUB, (NAZIONE), POS',
            'Foglio "Campi" (opzionale) CATEGORIA→CAMPI; default=2',
            'Ogni categoria ha CAMPI dedicati per batch di slot',
            'U12 non arbitra; evitare club/nazione ove possibile (Head Referee designato)',
            TIEBREAK_ORDER_NOTE,
            'In U12/U16/U20/Women il margine massimo conteggiato è 5-0',
            'Nessun "grace period" 15’; il mancato completamento senza permesso è forfeit (0-3)',
            'Arbitri per U12/Beginner sono assegnati dallo staff/club e non dai giocatori'
        ]
    })
    write_df_at(ws, note, 1, 1)

    ws2 = wb.create_sheet('Categorie_Riepilogo')
    rows = []
    for cat, packs in per_cat_output.items():
        rows.append({'Categoria': cat, 'Giocatori': packs['meta']['N'], 
                     'Gironi': packs['meta']['G'], 'Campi assegnati': packs['meta']['campi']})
    write_df_at(ws2, pd.DataFrame(rows), 1, 1)

    ws3 = wb.create_sheet('Campi_per_Categoria')
    write_df_at(ws3, pd.DataFrame([{'Categoria': k, 'Campi': v} for k, v in campi_map.items()]), 1, 1)

    for cat, packs in per_cat_output.items():
        write_calendar_and_classifica_sheet(wb, cat, packs['cal'], packs['classifica'])
        
        # Altri fogli informativi
        ws_g = wb.create_sheet(f"Gironi_{cat}"[:31])
        write_df_at(ws_g, packs['gironi'], 1, 1)
        
        ws_c = wb.create_sheet(f"Conformita_{cat}"[:31])
        write_df_at(ws_c, packs['conf'], 1, 1)
        
        ws_ra = wb.create_sheet(f"Riepilogo_Arbitri_{cat}"[:31])
        write_df_at(ws_ra, packs['rieparb'], 1, 1)
        
        ws_as = wb.create_sheet(f"Arbitri_per_Slot_{cat}"[:31])
        write_df_at(ws_as, packs['arb_slot'], 1, 1)

    out = safe_save_workbook(wb, path)
    _log(f"Output scritto: {out}")
    return out

# ------------------------------
# Category pipeline
# ------------------------------
def process_category(cat_name: str, df_players: pd.DataFrame, campi: int, random_seed: int = RANDOM_SEED):
    N = len(df_players)
    if N < 2:
        if PLACEHOLDER_MODE:
            _log(f"[{cat_name}] Categoria con pochi/no giocatori: placeholder.")
            df_gironi = pd.DataFrame([{'Girone': '-', 'Seed': '', 'POS': '', 'Giocatore': 'Nessun giocatore registrato', 'Club': '', 'Nazione': ''}])
            df_cal = pd.DataFrame([{'Girone': '-', 'Giornata': '-', 'Match #': '-', 'Casa': '-', 'Trasferta': '-', 'Campo': '-', 'Slot': '-', 'Gol Casa': '', 'Gol Trasferta': '', 'Arbitro': '-'}])
            df_class = pd.DataFrame([{'Girone': '-', 'Pos': 1, 'Giocatore': '-', 'Punti': 0, 'GF': 0, 'GS': 0, 'DG': 0}])
            df_conf = pd.DataFrame([{'Girone': '-', 'Conflitti CLUB': 0, 'Min CLUB (cat)': 0, 'Conflitti NAZIONE': 0, 'Min NAZIONE (cat)': 0, 'Raggiunti min teorici?': 'Sì'}])
            df_rieparb = pd.DataFrame([{'Giocatore': '-', 'Arbitraggi': 0}])
            df_arbslot = pd.DataFrame([{'Giornata': '-', 'Slot': '-', 'Arbitro': '-', 'Arbitraggi_slot': 0}])
            return {'gironi': df_gironi, 'cal': df_cal, 'classifica': df_class, 'conf': df_conf, 
                    'rieparb': df_rieparb, 'arb_slot': df_arbslot, 
                    'meta': {'N': N, 'G': 0, 'campi': campi}}
        else:
            _log(f"[{cat_name}] Giocatori insufficienti: ignorata.")
            return None

    sizes = compute_group_sizes(N)
    sizes = sorted(sizes)

    groups = draw_groups_fistf(df_players, sizes, random_seed=random_seed, avoid_fields=("CLUB", "NAZIONE"))

    labels = make_group_labels(len(sizes))
    seed_pos = set(df_players.sort_values('POS').iloc[:len(labels)]['POS'].tolist())
    
    rows = []
    for L in labels:
        for p in sorted(groups[L], key=lambda r: r['POS']):
            rows.append({'Girone': L, 'Seed': 'S' if p['POS'] in seed_pos else '', 
                         'POS': p['POS'], 'Giocatore': p['NOME'], 
                         'Club': p['CLUB'], 'Nazione': p.get('NAZIONE', '')})
    df_gironi = pd.DataFrame(rows).sort_values(['Girone', 'POS']).reset_index(drop=True)

    df_cal = build_fixtures_with_fields_and_slots(groups, campi, df_gironi)
    df_cal = assign_referees_by_slot(df_cal, df_gironi, category_name=cat_name)

    players_rows = []
    for L in labels:
        for p in sorted(groups[L], key=lambda r: r['POS']):
            players_rows.append({'Girone': L, 'Giocatore': p['NOME']})
    df_class_seed = pd.DataFrame(players_rows)

    df_conf = compute_draw_conformity(df_gironi)

    if not df_cal.empty:
        df_rieparb = df_cal['Arbitro'].value_counts().rename_axis('Giocatore').reset_index(name='Arbitraggi')
        df_arbslot = df_cal.groupby(['Giornata', 'Slot', 'Arbitro']).size().reset_index(name='Arbitraggi_slot')
    else:
        df_rieparb = pd.DataFrame([{'Giocatore': '-', 'Arbitraggi': 0}])
        df_arbslot = pd.DataFrame([{'Giornata': '-', 'Slot': '-', 'Arbitro': '-', 'Arbitraggi_slot': 0}])

    return {
        'gironi': df_gironi, 'cal': df_cal, 'classifica': df_class_seed,
        'conf': df_conf, 'rieparb': df_rieparb, 'arb_slot': df_arbslot,
        'meta': {'N': N, 'G': len(sizes), 'campi': campi}
    }

# ------------------------------
# FUNZIONE PRINCIPALE - Questa è quella chiamata dalla GUI
# ------------------------------
def run_v4d_full_export():
    """Funzione principale chiamata dalla GUI per generare il torneo"""
    log_debug("="*60)
    log_debug("FISTF ENGINE - INIZIO ELABORAZIONE")
    log_debug("="*60)
    
    # Ottieni percorso input da variabile d'ambiente
    input_file = os.environ.get('FISTF_CATEGORIE_PATH')
    
    if input_file and os.path.exists(input_file):
        log_debug(f"📂 File input: {input_file}")
        in_path = input_file
    else:
        log_debug("❌ ERRORE: File input non trovato!")
        raise FileNotFoundError('File input non trovato. Seleziona un file Categorie.xlsx valido.')
    
    # Ottieni percorso output da variabile d'ambiente
    output_path = os.environ.get('FISTF_OUTPUT_PATH')
    if not output_path:
        output_path = os.path.join(os.getcwd(), "FISTF_Tournament_Maker.xlsx")
    log_debug(f"📝 File output: {output_path}")
    
    # Carica Excel in sola lettura
    log_debug("📂 Caricamento file Excel...")
    try:
        xls = pd.ExcelFile(in_path, engine='openpyxl')
    except PermissionError:
        log_debug(f"❌ ERRORE: Il file è bloccato da Excel")
        raise PermissionError(f"Il file {os.path.basename(in_path)} è aperto in Excel. Chiudilo e riprova.")
    except Exception as e:
        log_debug(f"❌ ERRORE: {e}")
        raise
    
    sheets = [s for s in xls.sheet_names if s.strip()]
    log_debug(f"📊 Fogli trovati: {sheets}")
    
    campi_map = read_campi_mapping_if_any(xls, preset={})
    log_debug(f"📏 Campi per categoria: {campi_map}")
    
    cats = [s for s in sheets if s.strip().lower() not in ('campi', 'barrages')]
    log_debug(f"🏷️ Categorie da processare: {cats}")
    
    per_cat = {}
    
    for idx, cat in enumerate(cats, 1):
        log_debug(f"\n[{idx}/{len(cats)}] Processamento: {cat}")
        try:
            df_cat = load_players_sheet(xls, cat)
            log_debug(f"   ✅ Trovati {len(df_cat)} giocatori")
            campi = campi_map.get(cat, DEFAULT_CAMPI)
            pack = process_category(cat, df_cat, campi)
            if pack:
                per_cat[cat] = pack
                log_debug(f"   ✅ Completata (G={pack['meta']['G']})")
        except Exception as e:
            log_debug(f"   ❌ ERRORE: {e}")
            if PLACEHOLDER_MODE:
                campi = campi_map.get(cat, DEFAULT_CAMPI)
                per_cat[cat] = process_category(cat, pd.DataFrame(columns=['NOME','CLUB','NAZIONE','POS']), campi)
    
    if not per_cat:
        raise RuntimeError('Nessuna categoria processata.')
    
    log_debug(f"\n📊 Categorie processate: {list(per_cat.keys())}")
    log_debug("📝 Generazione file Excel...")
    
    final_path = export_multicat_excel(output_path, per_cat, campi_map)
    
    log_debug(f"✅ File generato: {final_path}")
    log_debug("="*60)
    log_debug("FISTF ENGINE - COMPLETATO")
    log_debug("="*60)
    
    return final_path

# ------------------------------
# MAIN per test diretto
# ------------------------------
if __name__ == '__main__':
    try:
        result = run_v4d_full_export()
        print(f"✅ File generato: {result}")
    except Exception as e:
        print(f"❌ Errore: {e}")
        traceback.print_exc()