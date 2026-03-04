# -*- coding: utf-8 -*-
"""
FISTF KO Engine (2–16 groups)
-----------------------------
Carica le formule ufficiali FISTF da `fistf_formulas_2_16.json` e genera:
- Barrages (se previsti)
- Round 1 (QF / R16 / R32)
- Turni successivi (R8 → R4 → R2) in automatico

Funzioni principali
- build_knockout_sheet(wb, source_ws_name, dest_ws_name, cls_col=12, start_row=1, json_path='fistf_formulas_2_16.json')

Requisiti
- openpyxl
- Il foglio sorgente `source_ws_name` deve contenere la Classifica di categoria con le colonne
  da `cls_col` in poi: L=Girone, M=Pos, N=Giocatore, O=Punti, P=GF, R=DG e colonna K con i KO-tag (1A, 2B, ...)

Nota
- Motore pensato per integrarsi con lo script v4d già in uso.
"""
from openpyxl.utils import get_column_letter

import json

def _sheet_quoted(name: str) -> str:
    return f"'{name}'" if any(ch in name for ch in " -()") else name


def _detect_last_row(src_ws, col_letter: str, start_row: int = 3) -> int:
    r = start_row
    while True:
        v = src_ws[f"{col_letter}{r}"].value
        if v is None or str(v).strip() in ('', '-'):
            break
        r += 1
    return r - 1


def _collect_groups(src_ws, col_group: str, first_row: int, last_row: int):
    groups = []
    for r in range(first_row, last_row + 1):
        g = src_ws[f"{col_group}{r}"].value
        if g not in (None, '', '-'):
            g = str(g).strip()
            if g not in groups:
                groups.append(g)
    return groups


def _token_to_player_formula(source_ws_name: str, token_cell_ref: str, rngK: str, rngN: str) -> str:
    # Restituisce formula Excel che risolve il token (es. "1A") in Giocatore (colonna N)
    return f"=IFERROR(INDEX({rngN}, MATCH({token_cell_ref}, {rngK}, 0)), \"\")"


def _write_header(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title)
    row += 1
    headers = ["Match", "Player 1", "Gol1", "Player 2", "Gol2", "Winner"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=row, column=j, value=h)
    return row + 1


def _winner_formula(ws, row: int) -> None:
    g1 = ws.cell(row=row, column=3).coordinate
    g2 = ws.cell(row=row, column=5).coordinate
    p1 = ws.cell(row=row, column=2).coordinate
    p2 = ws.cell(row=row, column=4).coordinate
    ws.cell(row=row, column=6, value=f"=IF(AND(ISNUMBER({g1}),ISNUMBER({g2})), IF({g1}>{g2},{p1}, IF({g2}>{g1},{p2},\"\")), \"\")")


def build_knockout_sheet(
    wb,
    source_ws_name: str,
    dest_ws_name: str,
    cls_col: int = 12,  # 12 -> L (Girone) nel foglio Classifica
    start_row: int = 1,
    json_path: str = 'fistf_formulas_2_16.json'
):
    """
    Genera (o rigenera) il foglio `dest_ws_name` contenente Barrages + KO Round + turni successivi
    usando le formule ufficiali FISTF caricate dal JSON per G=2..16.
    """
    if source_ws_name not in wb.sheetnames:
        raise ValueError(f"Foglio sorgente '{source_ws_name}' non trovato nel workbook.")

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            FORMULAS = json.load(f)
    except Exception as e:
        raise RuntimeError(f"Impossibile leggere il JSON delle formule: {e}")

    src = wb[source_ws_name]
    if dest_ws_name in wb.sheetnames:
        del wb[dest_ws_name]
    ws = wb.create_sheet(dest_ws_name)

    # Colonne in base a cls_col del foglio Classifica
    colL = get_column_letter(cls_col)       # Girone
    colN = get_column_letter(cls_col + 2)   # Giocatore

    # Individua ultima riga valida in colonna L
    last = _detect_last_row(src, colL, start_row=3)
    if last < 3:
        ws.cell(row=start_row, column=1, value="Nessun dato di classifica.")
        return ws

    # Determina numero gruppi (G)
    groups = _collect_groups(src, colL, 3, last)
    G = len(groups)
    if str(G) not in FORMULAS:
        raise ValueError(f"N. gruppi G={G} non supportato dal JSON {json_path}.")

    # Scrivi intestazione
    r = start_row
    ws.cell(row=r, column=1, value=f"FASE FINALE — Ufficiale FISTF — G={G}")
    r += 2

    # Range dal foglio sorgente (con nome foglio quotato)
    S = _sheet_quoted(source_ws_name)
    rngK = f"{S}!$K$3:{S}!$K${last}"
    rngN = f"{S}!${colN}$3:{S}!${colN}${last}"

    # Helper WIN/RUN
    ws.cell(row=r, column=1, value="Helper — WIN/RUN")
    r += 1
    for j,h in enumerate(["Type","KO","Giocatore","Valid"], start=1):
        ws.cell(row=r, column=j, value=h)
    r += 1

    # WIN rows (1A..GA)
    win_first = r
    for gidx in range(1, G+1):
        ws.cell(row=r, column=1, value="WIN")
        ws.cell(row=r, column=2, value=f'="{gidx}A"')
        ws.cell(row=r, column=3, value=_token_to_player_formula(source_ws_name, f"$B{r}", rngK, rngN))
        ws.cell(row=r, column=4, value=f'=IF($C{r}<>"",1,0)')
        r += 1
    win_last = r-1

    # RUN rows (1B..GB)
    run_first = r
    for gidx in range(1, G+1):
        ws.cell(row=r, column=1, value="RUN")
        ws.cell(row=r, column=2, value=f'="{gidx}B"')
        ws.cell(row=r, column=3, value=_token_to_player_formula(source_ws_name, f"$B{r}", rngK, rngN))
        ws.cell(row=r, column=4, value=f'=IF($C{r}<>"",1,0)')
        r += 1
    run_last = r-1
    r += 2

    def resolve_token(row_index: int, col_index: int, token: str, barrage_rows_map: dict):
        """Scrive in (row_index,col_index) il riferimento al giocatore per `token`.
        - Se `token` è Mx (winner barrage), aggancia alla colonna 6 della riga barrage corrispondente
        - Altrimenti risolve via INDEX/MATCH su rngK/rngN.
        """
        if token and isinstance(token, str) and token.upper().startswith('M') and token[1:].isdigit():
            m_id = token.upper()
            if m_id not in barrage_rows_map:
                raise KeyError(f"Riferimento a {m_id} non trovato (barrage non ancora scritto)")
            ws.cell(row=row_index, column=col_index, value=f"=$F{barrage_rows_map[m_id]}")
        else:
            # Scrivi il token in una colonna helper (H/I/J...) e usa INDEX/MATCH
            helper_col = 7 if col_index == 2 else 8  # G=7/H=8 come helper per p1/p2
            ws.cell(row=row_index, column=helper_col, value=token)
            ws.cell(row=row_index, column=col_index,
                    value=_token_to_player_formula(source_ws_name, ws.cell(row=row_index, column=helper_col).coordinate, rngK, rngN))

    # 1) BARRAGES (se previsti)
    formula = FORMULAS[str(G)]
    barrage_rows_map = {}
    if formula.get('barrages'):
        r = _write_header(ws, r, "BARRAGES — Ufficiale FISTF")
        for b in formula['barrages']:
            ws.cell(row=r, column=1, value=b.get('id', ''))
            # p1/p2
            resolve_token(r, 2, b.get('p1', ''), barrage_rows_map)
            resolve_token(r, 4, b.get('p2', ''), barrage_rows_map)
            _winner_formula(ws, r)
            barrage_rows_map[b.get('id','')] = r
            r += 1
        r += 1

    # 2) ROUND 1
    r = _write_header(ws, r, "ROUND 1")
    round1_rows = []
    for m in formula.get('round1', []):
        ws.cell(row=r, column=1, value=m.get('match',''))
        resolve_token(r, 2, m.get('p1',''), barrage_rows_map)
        resolve_token(r, 4, m.get('p2',''), barrage_rows_map)
        _winner_formula(ws, r)
        round1_rows.append(r)
        r += 1
    r += 1

    # 3) Turni successivi
    def write_next_round(title: str, prev_row_indexes: list, start_row_idx: int):
        rr = _write_header(ws, start_row_idx, title)
        out_rows = []
        for i in range(0, len(prev_row_indexes), 2):
            row_idx = rr
            ws.cell(row=row_idx, column=1, value=f"M{(i//2)+1}")
            ws.cell(row=row_idx, column=2, value=f"=$F{prev_row_indexes[i]}")
            ws.cell(row=row_idx, column=4, value=f"=$F{prev_row_indexes[i+1]}")
            _winner_formula(ws, row_idx)
            out_rows.append(row_idx)
            rr += 1
        return out_rows, rr + 1

    prev_rows = round1_rows
    for nr in formula.get('next_rounds', []):
        title_map = {
            'R32': 'TRENTADUESIMI (R32)',
            'R16': 'OTTAVI (R16)',
            'R8':  'QUARTI (R8)',
            'R4':  'SEMIFINALI (R4)',
            'R2':  'FINALE (R2)'
        }
        title = title_map.get(nr, nr)
        prev_rows, r = write_next_round(title, prev_rows, r)

    # Cosmesi colonne
    for col in ("A","B","C","D","E","F"):
        ws.column_dimensions[col].width = 28 if col in ("A","B","D","F") else 10

    return ws
